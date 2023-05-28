import pandas as pd 
from openpyxl import load_workbook #untuk berinterkasi antara python & excel file
#https://openpyxl.readthedocs.io/en/latest/api/openpyxl.worksheet.worksheet.html
from openpyxl.styles import *
from openpyxl.chart import *
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder 
from openpyxl.utils import get_column_letter
import json



input_file = 'input_data/supermarket_sales.xlsx'
output_file = 'output_data/report_penjualan_2019.xlsx'
configs = open('configs/webhook.json')
webhook_url = json.load(configs)['webhook_url']


##PART 1 - LOAD DATASET
df = pd.read_excel(input_file)

#Penjualan Total per Gender & Product Line
df = df.pivot_table(index=['Gender','Date'],
                    columns='Product line',
                    values='Total', 
                    aggfunc='sum').round()
                    

df.to_excel(output_file,
                sheet_name='Report',
                startrow=4)

print('Save dataframe done...')

#Select active worksheet
wb = load_workbook(output_file)
wb.active = wb['Report']

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row


# ##PART 2 - GRAFIK

barchart = BarChart()

data = Reference(wb.active, 
                min_col=min_column+2,
                max_col=max_column,
                min_row=min_row,
                max_row=max_row
                )

categories = Reference(wb.active,
                        min_col=min_column,
                        max_col=min_column,
                        min_row=min_row+1,
                        max_row=max_row
                        )

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)


wb.active.add_chart(barchart, 'J12')
barchart.title = 'Sales berdasarkan Produk'
barchart.style = 2

#Total
import string
alphabet = list(string.ascii_uppercase)
alphabet_excel = alphabet[:max_column]
#[A,B,C,D,E,F,G]
for i in alphabet_excel:
    if i != 'A':
        wb.active[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row})'
        wb.active[f'{i}{max_row+1}'].style = 'Currency'

wb.active[f'{alphabet_excel[0]}{max_row+1}'] = 'Total'

wb.active['A1'] = 'Sales Report'
wb.active['A2'] = '2019'
wb.active['A1'].font = Font('Arial', bold=True, size=20)
wb.active['A2'].font = Font('Arial', bold=True, size=10)

wb.save(output_file)

#PART - 3 Send to Discord
def send_to_discord():
    import discord
    from discord import SyncWebhook

    webhook = SyncWebhook.from_url(webhook_url)

    with open(file=output_file, mode='rb') as file:
        excel_file = discord.File(file)

    webhook.send('This is an automated report',
                 username='Sales Bot',
                 file=excel_file)

send_to_discord()





